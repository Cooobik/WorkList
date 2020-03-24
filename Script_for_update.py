import requests
from lxml import etree
from bs4 import BeautifulSoup
import re
import wget
import pandas
import xlrd
from mysql.connector import MySQLConnection, Error
import mysql.connector
from openpyxl import load_workbook
import pprint


mydb = mysql.connector.connect(
        host = "localhost",
        user = "root",
        password = "Вшьф_1999",
        database= "vpp_month"
    )


def insert_args(arg1, arg2):
    
    query1 = "INSERT INTO ua_npp_month_open(period,prodcom_cod,prodcome_name,unit_name,volume_month,volue_year_to_date) VALUES(%s,%s,%s,%s,%s,%s)"
    query2 = "INSERT INTO ua_exim_month_reg_open(period,prodcom_cod,region,prodcome_name,unit_name,volume_month,volue_year_to_date,volume_of_stocks) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)"
    
    cursor=mydb.cursor()
    try:
        cursor.executemany(query1, arg1)
        cursor.executemany(query2, arg2)
        mydb.commit()

    except Error as e:
        print('Error:', e)
    finally:
        cursor.close()
        mydb.close()


def del_files(arg1, arg2):        
    os.remove('C:/Users/Dima/'+ str(arg1)[-19:])
    os.remove('C:/Users/Dima/'+ str(arg2)[-20:])



def get_html(url):
    req = requests.get(url) 
    return req.text


def get_files(html):
    soup = BeautifulSoup(html, 'lxml')
    first = soup.find('a', href = re.compile('vpp_ue')).get('href')
    wget.download('http://www.ukrstat.gov.ua/' + str(first)[9:])
    second = soup.find('a', href = re.compile('vppr_ue')).get('href')
    wget.download('http://www.ukrstat.gov.ua/' + str(second)[9:])
    return first,second




def parse_files_num1(arg1):
    wb = load_workbook(filename = 'C:/Users/Dima/' + str(arg1)[-19:] )
    # wb = load_workbook(filename = 'C:/Users/Dima/vpp_ue_01_2020.xlsx')
    list_of_values = list()
    sheet = wb.active #['_T8']
    period = str(arg1)[-9:-5] + str(arg1)[-12:-10]
    # print(period)
    for index in range(5,sheet.max_row):
        if sheet.cell(row=index , column=6).value is None:
            break
        else:
            index_str = str(index)
            first_cell = 'A' + index_str
            last_cell = 'G' + index_str
            for row in sheet[first_cell:last_cell]:
                item = (period,
                    str(row[3].value),
                    str(row[0].value),
                    str(row[2].value).split(sep = ' / ')[0],
                    str(row[4].value).split(sep = ' / ')[0],
                    str(row[5].value).split(sep = ' / ')[0])
            list_of_values.append(item)        
    # print(list_of_values)
    return list_of_values




# def parse_files_num1(arg1):
#     wb = load_workbook(filename = 'C:/Users/Dima/' + str(arg1)[-19:] )
#     # wb = load_workbook(filename = 'C:/Users/Dima/vpp_ue_01_2020.xlsx')
#     list_of_values = list()
#     sheet = wb['_T8']
#     period = str(arg1)[-9:-5] + str(arg1)[-12:-10]
#     # print(period)
#     for index in range(5,sheet.max_row):
#         if sheet.cell(row=index , column=6).value is None:
#             break
#         else:
#             index_str = str(index)
#             first_cell = 'A' + index_str
#             last_cell = 'G' + index_str
#             for row in sheet[first_cell:last_cell]:
#                 row_data = [row[3],row[0],row[2],row[4],row[5]]
#                 string = period + ';'
#                 for cell in row_data:
#                     if cell.column == 1 or cell.column == 4 :
#                         string = string + str(cell.value) + ';'
#                     elif cell.column == 3 or cell.column == 5:
#                         s = str(cell.value)
#                         parts = s.split(sep=' / ')[0]
#                         string = string + parts +';'
#                     elif cell.column == 6:
#                         s = str(cell.value)
#                         parts = s.split(sep=' / ')[0]
#                         string = string + parts
#             list_of_values.append(string)        
#     # print(list_of_values)
#     return list_of_values


def parse_files_num2(arg2):
    wb = load_workbook(filename = 'C:/Users/Dima/' + str(arg2)[-20:] )
    # wb = load_workbook(filename = 'C:/Users/Dima/vppr_ue_01_2020.xlsx')
    sheet = wb.active #['_T7']
    period = str(arg2)[-9:-5] + str(arg2)[-12:-10]
    empty = ''
    list_of_strings = list()
    temp = []
    for index in range(5, sheet.max_row):
        index_str = str(index)
        first_cell = 'A' + index_str
        last_cell = 'H' + index_str
        for row in sheet[first_cell:last_cell]:
            if row[7].value is None:
                temp.clear()
                continue
            for cell in row:
                if cell.column >= 1 and cell.column <= 4:     
                    if cell.value != None:
                        temp.append(cell.value)
                    if cell.value == None:
                        cell.value = temp[cell.column - 1]
            item = (period,
                str(row[3].value),
                str(row[4].value).split(sep = ' / ')[0],
                str(row[0].value),
                str(row[2].value).split(sep = ' / ')[0],
                str(row[5].value).split(sep = ' / ')[0],
                empty,
                str(row[7].value).split(sep = ' / ')[0])
        list_of_strings.append(item)        
    # print(list_of_strings)
    return list_of_strings



# def parse_files_num2(arg2):
#     wb = load_workbook(filename = 'C:/Users/Dima/' + str(arg2)[-20:] )
#     # wb = load_workbook(filename = 'C:/Users/Dima/vppr_ue_01_2020.xlsx')
#     sheet = wb.active #['_T7']
#     period = str(arg2)[-9:-5] + str(arg2)[-12:-10]
#     list_of_strings = list()
#     temp = []
#     for index in range(5, sheet.max_row):
#         index_str = str(index)
#         first_cell = 'A' + index_str
#         last_cell = 'H' + index_str
#         for row in sheet[first_cell:last_cell]:
#             # print(first_cell, end=", ")
#             row_data = [row[3],row[4],row[0],row[2],row[5],row[6]]
#             if row[7].value is None:
#                 temp.clear()
#                 # print('')
#                 continue

#             string = period + ';'
#             for cell in row:
#                 if cell.column >= 1 and cell.column <= 4:     
#                     if cell.value != None:
#                         temp.append(cell.value)
#                     if cell.value == None:
#                         cell.value = temp[cell.column - 1]
#                 # print(cell.value, end=", ")
#             for cell in row_data:
#                 if cell.column == 1 or cell.column == 4 :
#                     string = string + str(cell.value) + ';'
#                 elif cell.column == 3 or cell.column == 5 or cell.column == 6:
#                     s = str(cell.value)
#                     parts = s.split(sep=' / ')[0]
#                     string = string + parts +';'
#                 elif cell.column == 7:
#                     s = str(cell.value)
#                     parts = s.split(sep=' / ')[0]
#                     string = string + parts
#                 # if cell.column == 3 or cell.column == 5:
#                 #     s = str(cell.value)
#                 #     parts = s.split(sep=' / ')[0]
#                 #     # parts_str = ','.join(parts)
#                 #     string = string + parts +','
#                 # else:
#                 #     string = string + str(cell.value) + ','
#             # print(last_cell)
#             # print(string)
#         list_of_strings.append(string)        
#     # print(list_of_strings)
#     return list_of_strings

def main():
    url = 'http://www.ukrstat.gov.ua/Noviny/new2020/new2020_u/new_u_02.html'
    arg1, arg2 = get_files(get_html(url))
    parse_files_num1(arg1)
    parse_files_num2(arg2)
    # print(parse_files_num2(arg2))
    # print(parse_files_num1(arg1))
    insert_args(parse_files_num1(arg1), parse_files_num2(arg2))
    del_files(arg1, arg2)
    # insert_args(parse_files_num1(arg1))

if __name__ == '__main__':
    main()








    # def parse_files_num1(arg1):
#     wb = load_workbook(filename = 'C:/Users/Dima/' + str(arg1)[-19:] )
#     # wb = load_workbook(filename = 'C:/Users/Dima/vpp_ue_01_2020.xlsx')
#     sheet = wb['_T8']

#     for index in range(5, 10):
#         index_str = str(index)
#         first_cell = 'A' + index_str
#         last_cell = 'G' + index_str
#         for row in sheet[first_cell:last_cell]:
#             string = ''
#             for cell in row:
#                 if cell.column == 3:
#                     s = str(cell.value)
#                     parts = s.split(sep=' / ')
#                     parts_str = ','.join(parts)
#                     # string = string + str(parts) +','
#                     string = string + parts_str +','
#                 else:
#                     string = string + str(cell.value) + ','
#              # print(string)





# def check_not_empty():
#     wb = load_workbook(filename = 'C:/Users/Dima/vppr_ue_01_2020.xlsx')
#     sheet = wb['_T7']
#     for index in range(5, 30):
#         index_str = str(index)
#         first_cell = 'A' + index_str
#         last_cell = 'H' + index_str
#         for column in sheet[first_cell:last_cell]:   
#             for cell in column:  
#                 if cell.value != None:
#                     temp_value = cell.value
#                 if cell.value == None:
#                     cell.value = temp_value
#                     print(cell.value)


# def check_not_empty():
#     wb = load_workbook(filename = 'C:/Users/Dima/vppr_ue_01_2020.xlsx')
#     sheet = wb['_T7']
#     for index in range(5, 30):
#         index_str = str(index)
#         first_cell = 'A' + index_str
#         last_cell = 'A' + index_str
#         for column in sheet[first_cell:last_cell]:   
#             for cell in column:  
#                 if cell.value != None:
#                     temp_value = cell.value
#                 if cell.value == None:
#                     cell.value = temp_value
#                     print(cell.value)





# def check_not_empty():
#     wb = load_workbook(filename = 'C:/Users/Dima/vppr_ue_01_2020.xlsx')
#     sheet = wb['_T7']
#     temp = []
#     for index in range(5, 20):
#         index_str = str(index)
#         first_cell = 'A' + index_str
#         last_cell = 'H' + index_str
#         for row in sheet[first_cell:last_cell]:
#             print(first_cell, end=", ")
            
#             if row[7].value is None:
#                 temp.clear()
#                 print('')
#                 continue

#             for cell in row:
#                 if cell.column >= 1 and cell.column <= 4:     
#                     if cell.value != None:
#                         temp.append(cell.value)
#                     if cell.value == None:
#                         cell.value = temp[cell.column - 1]
#                 print(cell.value, end=", ")
#             print(last_cell)