import requests
from lxml import etree
from bs4 import BeautifulSoup
import re
import wget
from mysql.connector import MySQLConnection, Error
import mysql.connector
from openpyxl import load_workbook
import pprint
from pyunpack import Archive
import os

mydb = mysql.connector.connect(
        host = "localhost",
        user = "root",
        password = "Вшьф_1999",
        database= "vpp_month"
    )


def insert_args(arg1, arg2):
    
    query1 = "INSERT INTO ua_npp_month_open(period,prodcom_cod,prodcome_name,unit_name,volume_month,volue_year_to_date) VALUES(%s,%s,%s,%s,%s,%s)"
    query2 = "INSERT INTO ua_exim_month_reg_open(period,prodcom_cod,region,prodcome_name,unit_name,volume_month,volue_year_to_date) VALUES(%s,%s,%s,%s,%s,%s,%s)"
    
    cursor=mydb.cursor()
    try:
        cursor.executemany(query1, arg1)
        cursor.executemany(query2, arg2)
        mydb.commit()

    except Error as e:
        print('Error:', e)
    finally:
        cursor.close()
        mydb.close()

def del_files(arg1, arg2):        
    os.remove('C:/Users/Dima/'+ str(arg1)[-16:])
    os.remove('C:/Users/Dima/'+ str(arg2)[-17:])


def get_html(url):
    req = requests.get(url) 
    return req.text


# def get_files(html):
#     soup = BeautifulSoup(html, 'lxml')
#     first = soup.find('a', href = re.compile(r'vpp_\w+\_u.xlsx')).get('href')
#     # wget.download('http://www.ukrstat.gov.ua/' + str(first)[9:])
#     second = soup.find('a', href = re.compile(r'vppr_\w+\_u.xlsx')).get('href')
#     # wget.download('http://www.ukrstat.gov.ua/' + str(second)[9:])
#     return first,second

def get_files(html):
    soup = BeautifulSoup(html, 'lxml')
    first = soup.find('a', href = re.compile(r'vpp_\w+\_u.xlsx')).get('href')
    wget.download('http://www.ukrstat.gov.ua/' + str(first)[9:], out='C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/')
    Archive('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/').extractall('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/') 

    second = soup.find('a', href = re.compile(r'vppr_\w+\_u.xlsx')).get('href')
    wget.download('http://www.ukrstat.gov.ua/' + str(second)[9:], out='C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/')
    Archive('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/').extractall('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/') 
    return first,second

def parse_files_num1(arg1):
    wb = load_workbook(filename = 'C:/Users/Dima/' + str(arg1)[-16:] )
    list_of_values = list()
    sheet = wb.active
    period = '20' + str(arg1)[-9:-7] + str(arg1)[-12:-10]
    for index in range(7,sheet.max_row):
        if sheet.cell(row=index , column=6).value is None:
            break
        else:
            index_str = str(index)
            first_cell = 'A' + index_str
            last_cell = 'H' + index_str
            for row in sheet[first_cell:last_cell]:
                item = (period,
                    str(row[2].value),
                    str(row[0].value),
                    str(row[1].value),
                    str(row[3].value),
                    str(row[4].value))
            list_of_values.append(item)        
    # print(list_of_values)
    return list_of_values



def parse_files_num2(arg2):
    wb = load_workbook(filename = 'C:/Users/Dima/' + str(arg2)[-17:] )
    sheet = wb.active
    period = '20' + str(arg2)[-9:-7] + str(arg2)[-12:-10]
    list_of_strings = list()
    temp1 = []
    temp2 = []
    for index in range(7, sheet.max_row):
        index_str = str(index)
        first_cell = 'A' + index_str
        last_cell = 'G' + index_str
        for row in sheet[first_cell:last_cell]:
            if row[0] and row[6] is None:
                temp1.clear()
                temp2.clear()
                continue
            if row[6].value is None:
                for cell in row:
                    if cell.column == 1:
                        temp1.append(cell.value)
                    if cell.column == 2:
                        temp2.append(cell.value)
                continue
            if row[6].value is not None:
                for cell in row:
                    if cell.column == 1: 
                        cell.value = cell.value + ', ' + temp1[cell.column - 1]
                        s = str(cell.value)
                        region_value = s.split(sep=', ')[0]
                        prodcom_name_value = s.split(sep=', ')[1]
                        unit_value = s.split(sep=', ')[2]        
                    if cell.column == 2:     
                        cell.value = temp2[cell.column - 2]   
            item = (period,
                str(row[1].value),
                region_value,
                prodcom_name_value,
                unit_value,
                str(row[2].value),
                str(row[3].value),
                str(row[6].value))
            list_of_strings.append(item)        
    # print(list_of_strings)
    return list_of_strings

def url_changer()
    list_of_urls = list()
    url ='http://www.ukrstat.gov.ua/Noviny/new2018/new2018_u/new_u01.html'
    
# def get_links_and_unzip():
#     soup = BeautifulSoup(html, 'lxml')
#     url ='http://www.ukrstat.gov.ua/Noviny/new2018/new2018_u/new_u01.html'
#     for i in range(1,len(url)):
#         try:
#             first = soup.find('a', href = re.compile(r'vpp_\w+\_u.xlsx')).get('href')
#             wget.download('http://www.ukrstat.gov.ua/' + str(first)[9:], out='C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/')
#             Archive('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/'+ filename[0].text).extractall('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/') 
#             second = soup.find('a', href = re.compile(r'vppr_\w+\_u.xlsx')).get('href')
#             wget.download('http://www.ukrstat.gov.ua/' + str(second)[9:], out='C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/')
#             Archive('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/'+ filename[0].text).extractall('C:/ProgramData/MySQL/MySQL Server 8.0/Data/vpp_month/')                        
#         except ValueError as ve:
#             print('ValueError Raised:', ve)
#     # return first,second  
#     return links

def main():
    url = 'http://www.ukrstat.gov.ua/Noviny/new2019/new2019_u/new_u_12.html'

    arg1, arg2 = get_files(get_html(url))
    parse_files_num1(arg1)
    parse_files_num2(arg2)
    # print(parse_files_num2(arg2))
    # print(parse_files_num1(arg1))
    # insert_args(parse_files_num1(arg1), parse_files_num2(arg2))
    # insert_args(parse_files_num1(arg1))
    # del_files(arg1, arg2)

if __name__ == '__main__':
    main()
